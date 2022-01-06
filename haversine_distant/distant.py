from openpyxl import load_workbook
from openpyxl import Workbook
from haversine import haversine




result = Workbook()
res = result.active

myFile = load_workbook('distance.xlsx')
mF = myFile.active

lat = mF['A']
lng = mF['B']
clat = mF['D']
clng = mF['E']

for i in range(len(lat) - 1):
    start = (lat[i + 1].value, lng[i + 1].value)
    goal = (clat[i + 1].value, clng[i + 1].value)
    dist = haversine(start, goal, unit = 'm')
    print(start, goal, dist)
    res.append([str(round(dist, 2)) + 'm'])
    # print(lat[i + 1].value, lng[i + 1].value, clat[i + 1].value, clng[i + 1].value)

result.save('result_1126.xlsx')