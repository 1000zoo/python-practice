from openpyxl import load_workbook
from openpyxl import Workbook
from haversine import haversine

class Point:
    def __init__ (self, latitude, longitude):
        self.latitude = latitude
        self.longitude = longitude

result = Workbook()
res = result.active

myFile = load_workbook('distance1112.xlsx')
mF = myFile.active

lat = mF['A']
lng = mF['B']

pointStack = []

for i in range(len(lat) - 2):
    start = Point(lat[i + 1].value, lng[i + 1].value)
    goal = Point(lat[i + 2].value, lng[i + 2].value)
    current = start
    pointStack.append(goal)

    while not not pointStack:
        temp = pointStack[-1]
        s = (current.latitude, current.longitude)
        g = (temp.latitude, temp.longitude)
        if haversine(s, g, unit='m') >= 0.995:
            midLat = (current.latitude + temp.latitude) / 2
            midLon = (current.longitude + temp.longitude) / 2
            currentMidpoint = Point(midLat, midLon)
            pointStack.append(currentMidpoint)
        else:
            print(current.latitude, current.longitude)
            res.append([current.latitude, current.longitude])
            # midPoints.append(current)
            current = pointStack.pop()

result.save('result_1126.xlsx')