from openpyxl import load_workbook
from openpyxl import Workbook
from haversine import haversine

class Point:
    def __init__ (self, latitude, longitude):
        self.latitude = latitude
        self.longitude = longitude

result = Workbook()
res = result.active

myFile = load_workbook('distance.xlsx')
mF = myFile.active

lat = mF['A']
lng = mF['B']
clat = mF['D']
clng = mF['E']

pointStack = []

for i in range(len(lat) -3):
    start = Point(lat[i + 2].value, lng[i + 2].value)
    goal = Point(lat[i + 3].value, lng[i + 3].value)
    current = start
    pointStack.append(goal)

    while not not pointStack:
        temp = pointStack[-1]
        s = (current.latitude, current.longitude)
        g = (temp.latitude, temp.longitude)
        if haversine(s, g, unit='m') >= 0.995:
            midLat = (current.latitude + temp.latitude) / 2
            midLon = (current.longitude + temp.longitude) / 2
            currentMidpoint = Point(midLat, midLon)
            pointStack.append(currentMidpoint)
        else:
            print(current.latitude, current.longitude)
            res.append([current.latitude, current.longitude])
            # midPoints.append(current)
            current = pointStack.pop()
    # dist = haversine(start, goal, unit = 'm')
    # print(start, goal, dist)
    # res.append([str(round(dist, 2)) + 'm'])

res.append(['c'])

for i in range(len(lat) -3):
    start = Point(clat[i + 2].value, clng[i + 2].value)
    goal = Point(clat[i + 3].value, clng[i + 3].value)
    current = start
    pointStack.append(goal)

    while not not pointStack:
        temp = pointStack[-1]
        s = (current.latitude, current.longitude)
        g = (temp.latitude, temp.longitude)
        if haversine(s, g, unit='m') >= 0.995:
            midLat = (current.latitude + temp.latitude) / 2
            midLon = (current.longitude + temp.longitude) / 2
            currentMidpoint = Point(midLat, midLon)
            pointStack.append(currentMidpoint)
        else:
            print(current.latitude, current.longitude)
            res.append([current.latitude, current.longitude])
            # midPoints.append(current)
            current = pointStack.pop()
result.save('result_1126.xlsx')