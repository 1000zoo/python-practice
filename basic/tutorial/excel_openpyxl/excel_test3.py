from openpyxl import load_workbook
from openpyxl import Workbook

new = Workbook()
k = new.active

i = 8

k.append([1,2,3,4,5])
k.append(['가', '나', '다'])
k.append(['a', 'b', 'sss'])
k.append(str(5) + 'm')

new.save('file.xlsx')

load_ex = load_workbook('file.xlsx', data_only=True)
temp = load_ex['Sheet']

print(temp.cell(1, 1).value + temp.cell(1, 2).value)

for c in temp.columns:
    print(c)